from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

OUT_PATH = Path(__file__).parent / "업무기록_템플릿.xlsx"

HEADERS = ["날짜", "업무종류", "소요시간(분)", "반복횟수", "애매했던점"]
TASK_TYPES = ["분류", "중복조회", "규격검토"]

FONT_NAME = "맑은 고딕"

wb = Workbook()
ws = wb.active
ws.title = "업무기록"

header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4472C4")
body_font = Font(name=FONT_NAME)

for col, title in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col, value=title)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

widths = [12, 12, 14, 10, 40]
for col, width in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + col)].width = width

for row in range(2, 32):
    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=row, column=col).font = body_font

dv = DataValidation(type="list", formula1=f'"{",".join(TASK_TYPES)}"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"B2:B31")

ws.freeze_panes = "A2"

wb.save(OUT_PATH)
print(f"저장됨: {OUT_PATH}")
